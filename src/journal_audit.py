# -*- coding: utf-8 -*-
"""
journal_audit.py — Journalisation des erreurs et anomalies du pipeline IJSS
===========================================================================
Se branche sur reconstruction_IJSS_v4_detail_unique.py sans en changer la logique.

Il produit DEUX sorties :
  1) un fichier .log horodaté (toutes sociétés confondues) — trace lisible ;
  2) un onglet Excel « 3 - Journal anomalies » PAR société — restitution auditable,
     avec code, gravité, salarié, contexte, explication et correction.

Chaque anomalie porte :
  - un CODE stable — soit un code du référentiel de divergence DSN (ERR/WRN/ALT-x-xxxxx-xxx,
    cf. "code de divergence - ERR V2.3 (3).csv"), soit un code interne ANO-0XX pour les
    quelques anomalies encore dormantes (jamais déclenchées par ce pipeline) ;
  - une GRAVITÉ (🔴 bloquant / 🟠 arbitrage / 🟡 à vérifier / ℹ️ info / ✅ corrigé) ;
  - une EXPLICATION (pourquoi c'est signalé) ;
  - un TYPE DE CORRECTION : AUTO (le script corrige) ou SIGNALEMENT (action humaine) ;
  - un LIEN vers la fiche du référentiel de divergence, quand le code en dispose.

Usage minimal :
    from journal_audit import JournalAudit, ANO
    J = JournalAudit(chemin_log="journal_IJSS.log")
    ...
    J.societe("AURA")
    J.tracer(ANO.ANNULATION_PURE, salarie="POUYET Marc", contexte="motif maladie",
             detail="aucun dépôt positif ne subsiste sur le trimestre")
    ...
    J.ajouter_onglet(wb)     # avant wb.save(...)
    J.cloturer()             # flush + résumé
"""

import logging
from datetime import datetime

# ===============================================================
# CATALOGUE DES ANOMALIES
#   code -> (titre, gravité, explication générique, type de correction)
#   gravités : "BLOQUANT" 🔴 | "ARBITRAGE" 🟠 | "VERIF" 🟡 | "INFO" ℹ️ | "CORRIGE" ✅
# ===============================================================
class ANO:
    """Codes d'anomalie du pipeline.

    Depuis l'ajout des codes manquants au référentiel de divergence (sous-catégories
    internes G2S Cortex 99001/99002, et nouveaux scénarios sur 99004/99005/99009/60001),
    la quasi-totalité des anomalies réellement déclenchées par ce pipeline sont désormais
    rattachées à un code du référentiel externe (ERR-x-xxxxx-xxx / WRN-x-xxxxx-xxx /
    ALT-x-xxxxx-xxx, cf. "code de divergence - ERR V2.3 (3).csv"). Leur lien de suivi
    (colonne "Liens" du CSV) est repris dans CATALOGUE et exposé dans la colonne "Lien"
    du journal.

    Seules restent en ANO-0XX les anomalies définies dans le catalogue mais jamais
    déclenchées par le pipeline actuel (dormantes, réservées à une évolution future) :
    FINPREV_MANQUANTE, ANNULATION_ORPHELINE, CHEVAUCHEMENT, EPISODE_ABSORBE,
    ARBITRAGE_MATERNITE, MOTIF_NON_MAPPE, STATUT_MANQUANT.
    """
    # --- Chargement / structure (bloquants) — sous-catégories internes G2S Cortex ---
    COLONNE_MANQUANTE   = "ERR-1-99002-001"   # Mapping colonnes DSN — colonne obligatoire introuvable
    FICHIER_SOURCE      = "ERR-1-99001-001"   # Fichier source — illisible ou introuvable
    SOCIETE_VIDE        = "WRN-2-99005-003"   # Nom usuel de la société — aucune ligne sur la période
    # --- Dates / cohérence de ligne ---
    DATE_INVALIDE_DJT   = "ERR-1-60002-002"   # DernierJour — format de date invalide
    DATE_INVALIDE_MOIS  = "ERR-1-99010-002"   # mois_absence — format invalide
    REPRISE_INCOHERENTE = "ERR-1-60010-002"   # RepriseDate — antérieure au dernier jour
    FINPREV_MANQUANTE   = "ANO-022"           # dormant — pas de code de divergence dédié
    # --- Annulations (qualifications calculées par la survivance) ---
    ANNULATION_DETECTEE = "ALT-2-99004-001"   # flag_annulation_G2SA — union de 3 signaux
    ANNULATION_PURE     = "ERR-2-99004-002"   # flag_annulation_G2SA — annulation pure
    ANNULATION_REDUITE  = "WRN-2-99004-003"   # flag_annulation_G2SA — durée réduite
    ANNULATION_AVERIF   = "WRN-2-99004-004"   # flag_annulation_G2SA — corrections multiples
    ANNULATION_ORPHELINE= "ANO-104"           # dormant — pas de code de divergence dédié
    # --- Chevauchement ---
    CHEVAUCHEMENT       = "ANO-200"           # dormant — pas de code de divergence dédié
    MEME_DJT            = "ALT-2-60001-002"   # Motif — chevauchement des motifs d'arrêt
    EPISODE_ABSORBE     = "ANO-202"           # dormant — pas de code de divergence dédié
    ARBITRAGE_MATERNITE = "ANO-203"           # dormant — pas de code de divergence dédié
    TPT_CHEVAUCHEMENT   = "ALT-2-60001-002"   # Motif — chevauchement des motifs d'arrêt
    # --- Contrôles finaux (ne devraient jamais arriver) ---
    INVARIANT           = "ERR-2-99009-001"   # Nb_jours_absences_SS — invariant jours=carence+IJSS violé
    RECOUVREMENT        = "ERR-2-60001-002"   # Motif — recouvrement résiduel après déchevauchement
    # --- Paramétrage (dormants — pas de code de divergence dédié) ---
    MOTIF_NON_MAPPE     = "ANO-400"
    STATUT_MANQUANT     = "ANO-401"

## CATALOGUE : code -> (titre, gravité, explication générique, type de correction, lien)
#   Le "lien" pointe vers la fiche du référentiel de codes de divergence
#   (monitor-app-drab.vercel.app) quand le code correspond à une erreur de champ DSN
#   répertoriée dans ce référentiel ; il est vide ("") pour les anomalies ANO-0XX propres
#   à la logique de calcul de ce pipeline (sans équivalent dans le référentiel externe).
CATALOGUE = {
    ANO.COLONNE_MANQUANTE:   ("Colonne DSN introuvable", "BLOQUANT",
        "Une colonne obligatoire du mapping est absente : le traitement ne peut pas se faire.",
        "SIGNALEMENT — vérifier le libellé dans find_col() ou renommer la colonne source.",
        "https://monitor-app-drab.vercel.app/errors/ERR-1-99002-001"),
    ANO.FICHIER_SOURCE:      ("Fichier source illisible", "BLOQUANT",
        "Le CSV/Excel est introuvable ou l'encodage n'est pas reconnu.",
        "SIGNALEMENT — vérifier le chemin et l'encodage (utf-8-sig / latin-1).",
        "https://monitor-app-drab.vercel.app/errors/ERR-1-99001-001"),
    ANO.SOCIETE_VIDE:        ("Société sans ligne", "INFO",
        "Aucune ligne DSN pour cette société : elle est ignorée.",
        "AUTO — société sautée, aucun fichier produit.",
        "https://monitor-app-drab.vercel.app/errors/WRN-2-99005-003"),
    ANO.DATE_INVALIDE_DJT:   ("DernierJour — format de date invalide", "VERIF",
        "Le dernier jour travaillé (DJT) n'est pas convertible en date : la ligne est écartée du calcul.",
        "AUTO — ligne exclue ; corriger le format de date (JJ/MM/AAAA) à la source.",
        "https://monitor-app-drab.vercel.app/errors/ERR-1-60002-002"),
    ANO.DATE_INVALIDE_MOIS:  ("mois_absence — format invalide", "VERIF",
        "Le mois d'absence n'est pas convertible en date : la ligne est écartée du calcul.",
        "AUTO — ligne exclue ; corriger le format du mois d'absence à la source.",
        "https://monitor-app-drab.vercel.app/errors/ERR-1-99010-002"),
    ANO.REPRISE_INCOHERENTE: ("Reprise antérieure ou égale au DJT", "VERIF",
        "La date de reprise est <= au dernier jour travaillé : incohérence de saisie DSN.",
        "AUTO — reprise ignorée pour le bornage ; à confirmer côté RH.",
        "https://monitor-app-drab.vercel.app/errors/ERR-1-60010-002"),
    ANO.FINPREV_MANQUANTE:   ("Fin prévisionnelle manquante", "VERIF",
        "Un épisode vivant n'a pas de date de fin prévisionnelle exploitable.",
        "AUTO — bornage sur la reprise si dispo, sinon épisode non ventilé ; à vérifier.",
        ""),
    ANO.ANNULATION_DETECTEE: ("Annulation détectée", "INFO",
        "Ligne marquée annulation (jours<0 OU motif=annulation OU flag_G2SA).",
        "AUTO — conservée et tracée ; ne supprime rien silencieusement.",
        "https://monitor-app-drab.vercel.app/errors/ALT-2-99004-001"),
    ANO.ANNULATION_PURE:     ("Annulation pure — IJSS à récupérer", "BLOQUANT",
        "Aucun dépôt positif ne subsiste : l'arrêt est entièrement annulé.",
        "SIGNALEMENT — vérifier RH/paie : suppression légitime ou IJSS à récupérer.",
        "https://monitor-app-drab.vercel.app/errors/ERR-2-99004-002"),
    ANO.ANNULATION_REDUITE:  ("Correction à durée réduite", "ARBITRAGE",
        "Un DJT initial est mort, un autre survit : arrêt re-déclaré à durée/début différents.",
        "AUTO — durée survivante retenue ; vérifier reprise anticipée (rubrique paie 3480).",
        "https://monitor-app-drab.vercel.app/errors/WRN-2-99004-003"),
    ANO.ANNULATION_AVERIF:   ("Corrections multiples à vérifier", "ARBITRAGE",
        "Plusieurs annulations ; certains mois annulés sans re-déclaration claire dans le trimestre.",
        "AUTO — épisode réel retenu ; vérifier qu'aucun mois n'est re-signalé ailleurs.",
        "https://monitor-app-drab.vercel.app/errors/WRN-2-99004-004"),
    ANO.ANNULATION_ORPHELINE:("Annulation non rattachée", "ARBITRAGE",
        "Une ligne 'annulation' n'a pas pu être rattachée à un motif réel (emp+DJT+mois).",
        "SIGNALEMENT — reste dans un groupe 'annulation' isolé ; vérifier le dépôt visé.",
        ""),
    ANO.CHEVAUCHEMENT:       ("Chevauchement tronqué", "INFO",
        "Deux motifs se chevauchent : le suivant est décalé (règle chronologique).",
        "AUTO — début du motif suivant décalé après le dernier jour absent du précédent.",
        ""),
    ANO.MEME_DJT:            ("Chevauchement des motifs d'arrêt", "ARBITRAGE",
        "Deux motifs d'arrêt se chevauchent au même dernier jour travaillé, ou un temps partiel "
        "thérapeutique chevauche un arrêt total à DJT différents : le déchevauchement automatique "
        "ne peut pas trancher seul dans tous les cas.",
        "SIGNALEMENT — arbitrage humain (grille de priorité à confirmer ; ex. MAT/PAT vs maladie, "
        "AT vs maladie, TPT vs arrêt total).",
        "https://monitor-app-drab.vercel.app/errors/ALT-2-60001-002"),
    ANO.EPISODE_ABSORBE:     ("Épisode entièrement recouvert", "VERIF",
        "Un épisode est intégralement couvert par un autre : 0 jour retenu.",
        "AUTO — épisode absorbé (0 j) et tracé ; vérifier s'il s'agit d'un doublon.",
        ""),
    ANO.ARBITRAGE_MATERNITE: ("Arbitrage maternité forcé", "INFO",
        "Date de début maternité forcée manuellement (ARBITRAGE_MATERNITE).",
        "AUTO — date imposée appliquée, épisodes voisins tronqués.",
        ""),
    ANO.INVARIANT:           ("Invariant jours = carence + IJSS violé", "BLOQUANT",
        "Sur une ligne mensuelle, jours d'absence != carence + IJSS : erreur de calcul.",
        "SIGNALEMENT — ne devrait jamais arriver ; investiguer le calcul de carence.",
        "https://monitor-app-drab.vercel.app/errors/ERR-2-99009-001"),
    ANO.RECOUVREMENT:        ("Recouvrement résiduel entre motifs", "BLOQUANT",
        "Après déchevauchement, deux motifs se recouvrent encore : bug de bornage.",
        "SIGNALEMENT — ne devrait jamais arriver ; investiguer le déchevauchement.",
        "https://monitor-app-drab.vercel.app/errors/ERR-2-60001-002"),
    ANO.MOTIF_NON_MAPPE:     ("Motif non paramétré", "VERIF",
        "Le motif n'a pas de code DSN / source juridique / taux dédié.",
        "AUTO — fallback (taux 50%, 'référence non définie') ; compléter le paramétrage.",
        ""),
    ANO.STATUT_MANQUANT:     ("Statut ou temps de travail manquant", "VERIF",
        "Statut RC ou modalité de temps absent : impacte le futur maintien de salaire.",
        "SIGNALEMENT — récupérer le statut (bloc S21.G00.40).",
        ""),
}

_EMOJI = {"BLOQUANT": "🔴", "ARBITRAGE": "🟠", "VERIF": "🟡", "INFO": "ℹ️", "CORRIGE": "✅"}
_NIVEAU_LOG = {"BLOQUANT": logging.ERROR, "ARBITRAGE": logging.WARNING,
               "VERIF": logging.WARNING, "INFO": logging.INFO, "CORRIGE": logging.INFO}


class JournalAudit:
    def __init__(self, chemin_log="journal_IJSS.log"):
        self.chemin_log = chemin_log
        self.entrees = []          # liste de dicts (pour l'onglet Excel)
        self._societe = "—"
        self._logger = logging.getLogger("journal_ijss")
        self._logger.setLevel(logging.INFO)
        self._logger.handlers.clear()
        fh = logging.FileHandler(chemin_log, mode="w", encoding="utf-8")
        fh.setFormatter(logging.Formatter("%(asctime)s | %(levelname)-7s | %(message)s",
                                          datefmt="%Y-%m-%d %H:%M:%S"))
        self._logger.addHandler(fh)
        ch = logging.StreamHandler()
        ch.setFormatter(logging.Formatter("%(message)s"))
        self._logger.addHandler(ch)
        self._logger.info(f"=== Démarrage journal IJSS — {datetime.now():%Y-%m-%d %H:%M:%S} ===")

    def societe(self, nom):
        """Étiquette les anomalies suivantes avec cette société."""
        self._societe = str(nom)
        self._logger.info(f"----- Société : {self._societe} -----")

    def tracer(self, code, salarie="", contexte="", detail="", correction_auto=None):
        """Enregistre une anomalie du catalogue.
        - detail : précision libre sur le cas rencontré ;
        - correction_auto : si le script a appliqué une correction, passer son libellé
          (bascule la gravité affichée en ✅ « corrigé »)."""
        titre, gravite, explication, correction, lien = CATALOGUE.get(
            code, ("Anomalie non cataloguée", "VERIF", detail, "SIGNALEMENT", ""))
        grav_aff = "CORRIGE" if correction_auto else gravite
        message = explication if not detail else f"{explication} — {detail}"
        corr_txt = correction_auto if correction_auto else correction
        self.entrees.append({
            "Code": code, "Gravité": _EMOJI[grav_aff], "Société": self._societe,
            "Salarié": salarie, "Contexte": contexte, "Titre": titre,
            "Explication": message, "Correction": corr_txt, "Lien": lien,
        })
        lien_txt = f" | {lien}" if lien else ""
        self._logger.log(_NIVEAU_LOG[grav_aff],
                         f"{code} {_EMOJI[grav_aff]} [{self._societe}] "
                         f"{salarie or '—'} | {titre} | {message} | {corr_txt}{lien_txt}")

    def erreur_libre(self, message):
        """Erreur inattendue (exception) hors catalogue."""
        self.entrees.append({
            "Code": "ANO-999", "Gravité": "🔴", "Société": self._societe,
            "Salarié": "", "Contexte": "exception", "Titre": "Erreur inattendue",
            "Explication": str(message), "Correction": "SIGNALEMENT — investiguer.", "Lien": "",
        })
        self._logger.error(f"ANO-999 🔴 [{self._societe}] {message}")

    def resume(self):
        """Compte par code, pour le log final."""
        cpt = {}
        for e in self.entrees:
            cpt[e["Code"]] = cpt.get(e["Code"], 0) + 1
        return dict(sorted(cpt.items()))

    def ajouter_onglet(self, wb, titre="3 - Journal anomalies"):
        """Ajoute au classeur openpyxl un onglet listant les anomalies de la société
        courante (ou toutes si société non filtrée). Style aligné sur le pipeline."""
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        NAVY = "1F3864"; REDF = "F4D6D2"; ORANGEF = "FBE7CE"; YEL = "FEF3CD"
        GREENF = "DDEBE2"; LIGHT = "F2F4F7"
        lignes = [e for e in self.entrees if e["Société"] in (self._societe, "—")] \
                 or self.entrees
        ws = wb.create_sheet(titre)
        cols = ["Code", "Gravité", "Salarié", "Contexte", "Titre", "Explication", "Correction", "Lien"]
        thin = Side(style="thin", color="D0D5DD")
        bd = Border(left=thin, right=thin, top=thin, bottom=thin)
        for j, c in enumerate(cols, 1):
            cell = ws.cell(1, j, c)
            cell.font = Font(bold=True, size=9, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor=NAVY)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = bd
        for i, e in enumerate(lignes, start=2):
            for j, c in enumerate(cols, 1):
                val = e.get(c, "")
                cell = ws.cell(i, j, val)
                cell.border = bd
                cell.alignment = Alignment(horizontal="left", vertical="center",
                                           wrap_text=(c in ("Explication", "Correction")))
                if c == "Lien" and val:
                    cell.hyperlink = val
                    cell.font = Font(size=9, color="0563C1", underline="single")
                else:
                    cell.font = Font(size=9)
            g = e["Gravité"]
            fill = REDF if g == "🔴" else ORANGEF if g == "🟠" else YEL if g == "🟡" \
                   else GREENF if g == "✅" else LIGHT
            for j in range(1, len(cols) + 1):
                ws.cell(i, j).fill = PatternFill("solid", fgColor=fill)
        larg = {"Code": 15, "Gravité": 8, "Salarié": 20, "Contexte": 22,
                "Titre": 30, "Explication": 55, "Correction": 55, "Lien": 42}
        for j, c in enumerate(cols, 1):
            ws.column_dimensions[get_column_letter(j)].width = larg[c]
        ws.freeze_panes = "A2"
        return ws

    def cloturer(self):
        r = self.resume()
        self._logger.info(f"=== Fin — {len(self.entrees)} anomalie(s) : {r} ===")
        for h in self._logger.handlers:
            h.flush()
