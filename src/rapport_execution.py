# -*- coding: utf-8 -*-
"""
rapport_execution.py — Rapport d'exécution cumulatif (n'écrase jamais)
======================================================================
À chaque run, AJOUTE :
  - une section datée dans un fichier Markdown (rapport_executions.md) ;
  - des lignes dans un classeur Excel (rapport_executions.xlsx) qui reprend
    TOUTES les informations du Markdown, réparties sur 4 onglets :
        1) Exécutions        — 1 ligne par run (synthèse)
        2) Détail sociétés   — 1 ligne par (run, société) : compteurs + familles
        3) Anomalies         — 1 ligne par (run, anomalie) : code, gravité, détail
        4) Sortie terminal   — 1 ligne par run : capture intégrale du terminal
L'Excel est ouvert en lecture puis ré-écrit (append fiable via openpyxl).

Usage :
    R = RapportExecution(md_path, xlsx_path, source="....csv")
    R.demarrer()
    ... prints ...
    R.ajouter_societe(res, fichier)      # res = dict renvoyé par traiter_societe
    R.finaliser(journal=J)
"""
import sys
import os
import io
from datetime import datetime

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

NAVY = "1F3864"; LIGHT = "F2F4F7"; REDF = "F4D6D2"; ORANGEF = "FBE7CE"
YEL = "FEF3CD"; GREENF = "DDEBE2"
_THIN = Side(style="thin", color="D0D5DD")
_BD = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

# Onglets et en-têtes
FEUILLES = {
    "1 - Exécutions": ["Horodatage", "Date", "Heure", "Durée (s)", "Fichier source",
                       "Sociétés", "Total positives", "Total annulations",
                       "🔴 Bloquant", "🟠 Arbitrage", "🟡 À vérifier", "ℹ️ Info", "✅ Corrigé",
                       "Total anomalies", "Fichiers produits"],
    "2 - Détail sociétés": ["Horodatage", "Société", "Lignes positives", "Annulations",
                            "Familles annul.", "COMPENSEE", "PURE", "REDUITE", "AVERIF",
                            "Déchevauchements", "Fichier produit"],
    "3 - Anomalies": ["Horodatage", "Code", "Gravité", "Société", "Salarié",
                      "Titre", "Explication", "Correction"],
    "4 - Sortie terminal": ["Horodatage", "Sortie terminal complète"],
}


class _Tee:
    """Duplique tout ce qui est écrit sur le flux réel ET dans un tampon mémoire."""
    def __init__(self, flux_reel):
        self._reel = flux_reel
        self.tampon = io.StringIO()

    def write(self, texte):
        self._reel.write(texte)
        self.tampon.write(texte)

    def flush(self):
        self._reel.flush()


class RapportExecution:
    def __init__(self, md_path, xlsx_path, source=""):
        self.md_path = md_path
        self.xlsx_path = xlsx_path
        self.source = source
        self.debut = None
        self.societes = []      # dicts : societe, n_pos, n_ann, cats, n_chevauchement, fichier
        self._tee = None
        self._stdout_origine = None

    def demarrer(self):
        self.debut = datetime.now()
        self._stdout_origine = sys.stdout
        self._tee = _Tee(sys.stdout)
        sys.stdout = self._tee

    def ajouter_societe(self, res, fichier):
        self.societes.append({
            "societe": res.get("societe", ""),
            "n_pos": res.get("n_pos", 0),
            "n_ann": res.get("n_ann", 0),
            "cats": res.get("cats", {}) or {},
            "n_familles": res.get("n_familles", 0),
            "n_chevauchement": len(res.get("journal", []) or []),
            "fichier": os.path.basename(fichier) if fichier else "",
        })

    def _restaurer_stdout(self):
        if self._stdout_origine is not None:
            sys.stdout = self._stdout_origine

    def _compter_anomalies(self, journal):
        c = {"🔴": 0, "🟠": 0, "🟡": 0, "ℹ️": 0, "✅": 0}
        entrees = getattr(journal, "entrees", []) if journal else []
        for e in entrees:
            g = e.get("Gravité", "")
            if g in c:
                c[g] += 1
        return c, entrees

    # ---------------- Excel (append multi-onglets) ----------------
    def _ouvrir_ou_creer_classeur(self):
        if os.path.exists(self.xlsx_path):
            wb = load_workbook(self.xlsx_path)
            # s'assurer que tous les onglets existent (compat versions antérieures)
            for titre, entetes in FEUILLES.items():
                if titre not in wb.sheetnames:
                    self._creer_feuille(wb, titre, entetes)
            return wb
        wb = Workbook()
        wb.remove(wb.active)
        for titre, entetes in FEUILLES.items():
            self._creer_feuille(wb, titre, entetes)
        return wb

    def _creer_feuille(self, wb, titre, entetes):
        ws = wb.create_sheet(titre)
        for j, h in enumerate(entetes, 1):
            c = ws.cell(1, j, h)
            c.font = Font(bold=True, size=9, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor=NAVY)
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = _BD
        ws.freeze_panes = "A2"
        return ws

    def _style_ligne(self, ws, ligne, valeurs, fill=None, wrap_cols=()):
        for j, v in enumerate(valeurs, 1):
            c = ws.cell(ligne, j, v)
            c.border = _BD
            c.font = Font(size=9)
            c.alignment = Alignment(horizontal="left", vertical="center",
                                    wrap_text=(j in wrap_cols))
            if fill:
                c.fill = PatternFill("solid", fgColor=fill)

    def finaliser(self, journal=None):
        fin = datetime.now()
        duree = (fin - self.debut).total_seconds() if self.debut else 0.0
        sortie_terminal = self._tee.tampon.getvalue() if self._tee else ""
        self._restaurer_stdout()

        cpt, entrees = self._compter_anomalies(journal)
        tot_pos = sum(s["n_pos"] for s in self.societes)
        tot_ann = sum(s["n_ann"] for s in self.societes)
        stamp = self.debut.strftime("%Y-%m-%d %H:%M:%S") if self.debut else ""
        d_jour = self.debut.strftime("%Y-%m-%d") if self.debut else ""
        d_heure = self.debut.strftime("%H:%M:%S") if self.debut else ""
        anomalies_reelles = [e for e in entrees if e.get("Gravité") != "ℹ️"]
        fichiers = " | ".join(s["fichier"] for s in self.societes)

        # ============ 1) MARKDOWN (append) ============
        self._ecrire_markdown(stamp, duree, tot_pos, tot_ann, cpt, entrees,
                              anomalies_reelles, sortie_terminal)

        # ============ 2) EXCEL (append multi-onglets) ============
        wb = self._ouvrir_ou_creer_classeur()

        ws1 = wb["1 - Exécutions"]
        r = ws1.max_row + 1
        self._style_ligne(ws1, r, [
            stamp, d_jour, d_heure, round(duree, 1), os.path.basename(self.source),
            len(self.societes), tot_pos, tot_ann,
            cpt["🔴"], cpt["🟠"], cpt["🟡"], cpt["ℹ️"], cpt["✅"],
            len(anomalies_reelles), fichiers,
        ], fill=(LIGHT if r % 2 else "FFFFFF"), wrap_cols=(15,))

        ws2 = wb["2 - Détail sociétés"]
        for s in self.societes:
            r = ws2.max_row + 1
            cats = s["cats"]
            self._style_ligne(ws2, r, [
                stamp, s["societe"], s["n_pos"], s["n_ann"], s["n_familles"],
                cats.get("COMPENSEE", 0), cats.get("PURE", 0),
                cats.get("REDUITE", 0), cats.get("AVERIF", 0),
                s["n_chevauchement"], s["fichier"],
            ], fill=(LIGHT if r % 2 else "FFFFFF"))

        ws3 = wb["3 - Anomalies"]
        for e in entrees:
            r = ws3.max_row + 1
            g = e.get("Gravité", "")
            fill = REDF if g == "🔴" else ORANGEF if g == "🟠" else YEL if g == "🟡" \
                   else GREENF if g == "✅" else LIGHT
            self._style_ligne(ws3, r, [
                stamp, e.get("Code", ""), g, e.get("Société", ""),
                e.get("Salarié", "") or "—", e.get("Titre", ""),
                e.get("Explication", ""), e.get("Correction", ""),
            ], fill=fill, wrap_cols=(7, 8))

        ws4 = wb["4 - Sortie terminal"]
        r = ws4.max_row + 1
        self._style_ligne(ws4, r, [stamp, sortie_terminal.rstrip()],
                          fill=(LIGHT if r % 2 else "FFFFFF"), wrap_cols=(2,))
        ws4.cell(r, 2).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

        # largeurs de colonnes
        largeurs = {
            "1 - Exécutions": {1: 20, 4: 9, 5: 42, 15: 60},
            "2 - Détail sociétés": {1: 20, 2: 12, 11: 34},
            "3 - Anomalies": {1: 20, 5: 20, 6: 30, 7: 55, 8: 55},
            "4 - Sortie terminal": {1: 20, 2: 120},
        }
        for titre, cfg in largeurs.items():
            ws = wb[titre]
            for col, w in cfg.items():
                ws.column_dimensions[get_column_letter(col)].width = w

        os.makedirs(os.path.dirname(self.xlsx_path), exist_ok=True)
        wb.save(self.xlsx_path)

        print(f"📝 Rapport ajouté : {self.md_path}")
        print(f"📊 Suivi Excel mis à jour : {self.xlsx_path}")

    def _ecrire_markdown(self, stamp, duree, tot_pos, tot_ann, cpt, entrees,
                         anomalies_reelles, sortie_terminal):
        premiere_fois = not os.path.exists(self.md_path)
        L = []
        if premiere_fois:
            L.append("# Rapport d'exécutions — Reconstruction IJSS\n")
            L.append("Historique cumulatif. Chaque exécution ajoute une section "
                     "ci-dessous (le fichier n'est jamais écrasé).\n")
        L.append(f"\n---\n\n## Exécution du {stamp}\n")
        L.append(f"- **Durée** : {duree:.1f} s")
        L.append(f"- **Fichier source** : `{os.path.basename(self.source)}`")
        L.append(f"- **Sociétés traitées** : {len(self.societes)}")
        L.append(f"- **Total lignes** : {tot_pos} positives · {tot_ann} annulations")
        L.append(f"- **Anomalies** : 🔴 {cpt['🔴']} · 🟠 {cpt['🟠']} · "
                 f"🟡 {cpt['🟡']} · ℹ️ {cpt['ℹ️']} · ✅ {cpt['✅']}\n")
        if self.societes:
            L.append("| Société | Lignes pos. | Annulations | Fichier produit |")
            L.append("| --- | --- | --- | --- |")
            for s in self.societes:
                L.append(f"| {s['societe']} | {s['n_pos']} | {s['n_ann']} | "
                         f"`{s['fichier']}` |")
            L.append("")
        if anomalies_reelles:
            L.append("### Anomalies signalées\n")
            L.append("| Code | Gravité | Société | Salarié | Titre |")
            L.append("| --- | --- | --- | --- | --- |")
            for e in anomalies_reelles:
                L.append(f"| {e['Code']} | {e['Gravité']} | {e['Société']} | "
                         f"{e['Salarié'] or '—'} | {e['Titre']} |")
            L.append("")
        L.append("<details><summary>Sortie terminal complète</summary>\n")
        L.append("```")
        L.append(sortie_terminal.rstrip())
        L.append("```")
        L.append("</details>\n")
        os.makedirs(os.path.dirname(self.md_path), exist_ok=True)
        with open(self.md_path, "a", encoding="utf-8") as f:
            f.write("\n".join(L) + "\n")
