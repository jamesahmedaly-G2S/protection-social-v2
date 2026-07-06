# -*- coding: utf-8 -*-
"""Écriture du classeur Excel par société."""
import os
import pandas as pd
from src.normalisation import _norm_soc
from config import OUTPUT_DIR

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

COLS = [
    "Nom", "Prénom", "NIR", "Matricule", "Nom usage", "Société", "Statut", "Temps de travail",
    "Type de ligne", "Motif arrêt", "Code DSN", "Source juridique", "DJT",
    "Date fin prévisionnelle", "Date reprise", "Mois DSN", "Jours absence (mois)",
    "Début carence", "Fin carence", "Jours carence (mois)", "Jours IJSS SS (mois)",
    "Jours SS CPAM (J1-J90)", "Jours Prévoyance J91+", "Total j abs Q1-2026", "Total IJSS Q1-2026",
    "J avant Q1-2026", "IJSS avant Q1-2026", "⭐ Total épisode complet", "⭐ Total IJSS épisode complet",
    "Date début prévoyance", "Solde net épisode (pos+neg)", "Statut prévoyance",
    "🔍 Interprétation épisode (auto)",
]
# groupes : (titre, première colonne 1-based, dernière colonne 1-based)
GROUPES = [("IDENTIFICATION", 1, 8), ("ARRÊT & MOTIF", 9, 15),
           ("MOIS COURANT", 16, 23), ("ÉPISODE COMPLET", 24, 33)]
NAVY="1F3864"; GREY="5B6B7F"; LIGHT="F2F4F7"; REDF="F4D6D2"; ORANGEF="FBE7CE"; GREENF="DDEBE2"


def ecrire_xlsx(res, J=None):
    detail = res["detail"]; soc = res["societe"]
    wb = Workbook(); ws = wb.active; ws.title = "1 - Détail épisodes"
    thin = Side(style="thin", color="D0D5DD")
    bd = Border(left=thin, right=thin, top=thin, bottom=thin)
    # rangée 1 : en-têtes colonnes (plus de bandeau ni d'en-têtes groupés au-dessus)
    for j, name in enumerate(COLS, start=1):
        c = ws.cell(1, j, name); c.font = Font(bold=True, size=9, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = bd
    ws.row_dimensions[1].height = 42
    # données
    r0 = 2
    for i, (_, row) in enumerate(detail.iterrows()):
        is_ann = row["Type de ligne"].startswith("❌")
        for j, name in enumerate(COLS, start=1):
            val = row.get(name, "")
            c = ws.cell(r0 + i, j, val); c.border = bd
            c.font = Font(size=9, color=("9B2C20" if is_ann else "222222"))
            c.alignment = Alignment(horizontal=("left" if j <= 15 or j == 32 or j == 33 else "center"),
                                    vertical="center", wrap_text=(j == 33))
        fill = REDF if is_ann else (LIGHT if i % 2 else "FFFFFF")
        for j in range(1, len(COLS) + 1):
            ws.cell(r0 + i, j).fill = PatternFill("solid", fgColor=fill)
        interp = str(row.get("🔍 Interprétation épisode (auto)", ""))
        if interp.startswith("✅"):
            ws.cell(r0 + i, len(COLS)).fill = PatternFill("solid", fgColor=GREENF)
        elif interp.startswith("🟠"):
            ws.cell(r0 + i, len(COLS)).fill = PatternFill("solid", fgColor=ORANGEF)
        elif interp.startswith("🔴"):
            ws.cell(r0 + i, len(COLS)).fill = PatternFill("solid", fgColor=REDF)
    ws.freeze_panes = "A2"
    # largeurs
    larg = {"Nom":16,"Prénom":12,"NIR":15,"Matricule":12,"Nom usage":12,"Société":12,"Statut":10,
            "Temps de travail":11,"Type de ligne":15,"Motif arrêt":26,"Code DSN":7,"Source juridique":22,
            "DJT":11,"Date fin prévisionnelle":12,"Date reprise":11,"Mois DSN":10,
            "🔍 Interprétation épisode (auto)":60,"Statut prévoyance":24,"Solde net épisode (pos+neg)":11}
    for j, name in enumerate(COLS, start=1):
        ws.column_dimensions[get_column_letter(j)].width = larg.get(name, 9)
    # onglet secondaire : journal chevauchements
    if res["journal"]:
        ws2 = wb.create_sheet("2 - Journal chevauchements")
        jdf = pd.DataFrame(res["journal"])
        for j, name in enumerate(jdf.columns, start=1):
            cc = ws2.cell(1, j, name); cc.font = Font(bold=True, color="FFFFFF")
            cc.fill = PatternFill("solid", fgColor=NAVY)
        for i, (_, rr) in enumerate(jdf.iterrows(), start=2):
            for j, name in enumerate(jdf.columns, start=1):
                ws2.cell(i, j, str(rr[name]))
        for j in range(1, len(jdf.columns) + 1):
            ws2.column_dimensions[get_column_letter(j)].width = 24
    if J is not None:
        J.societe(soc); J.ajouter_onglet(wb)
    nom = f"Reconstruit_{_norm_soc(soc).replace(' ', '-')}_CORRIGE.xlsx"
    out = os.path.join(OUTPUT_DIR, nom)
    wb.save(out)
    print(f"   ✅ écrit -> {out}  ({res['n_pos']} pos + {res['n_ann']} annul.)")
    return out

